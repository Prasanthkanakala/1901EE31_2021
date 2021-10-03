print('My name is Kanakala Durga Prasanth. My roll no is 1901ee31')

#importing modules
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

os.system('cls')

#created a map to convert the grade to score in numericals       
grade_to_score = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,
                'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}



def result_calculate(Spi,ws_sheet,Credits):   #Function to calculate final result    
    ws_sheet.append(["Semester No"]+[semno for semno in range(1,9)])
    #counting credits
    ws_sheet.append(["Semester wise Credits taken"]+Credits)
    ws_sheet.append(["SPI per sem"]+Spi)
    
    pre_credits_list=[]
    for i in range(len(Credits)):
        pre_credits_list.append(sum(Credits[:i+1]))

    ws_sheet.append(["Total Credits taken"]+pre_credits_list) #counting total credits
    CPI_num=[]
    for i in range(len(Credits)):
        CPI_num.append(Spi[i]*Credits[i])
    
    #CPI calculation formula
    CPI_Cal=[]
    for i in range(len(Spi)):
        CPI_Cal.append(round(sum(CPI_num[:i+1])/pre_credits_list[i],2)) 
    ws_sheet.append(["Total CPI "]+CPI_Cal)
    return

#function to make score sheets in output files
def score_sheet_maker(subject_dict,length):
    #opening and reading grades file
    with open("grades.csv", "r") as score_data:
        score_csv = csv.reader(score_data)
        #creating a list for score data
        
        score_list=[]
        for data in score_csv:
            score_list.append(list(data))
        score_list=score_list[1:]
        #initializing the counter
        counter=1
        for data in score_list:
            Rollno,Sem_no, = data[0],data[1]  #initializing the roll no
            file_path='./output/'+'{}.xlsx'.format(Rollno)
            if not os.path.isfile(file_path): #checking the exitance of a file
                get_workbook(data) #creating a new workbook
            #loading the workbook
            wb_sheet=load_workbook(r'output\\{}.xlsx'.format(Rollno))
            #checking the existance of sheet
            if not is_sheet_exist(f'Sem{Sem_no}',wb_sheet):
                wb_sheet.create_sheet(f'Sem{Sem_no}',int(Sem_no))
                ws_sheet = wb_sheet["Sem{}".format(int(Sem_no))]
                ws_sheet.column_dimensions["C"].width = length
            ws_sheet = wb_sheet[f"Sem{int(Sem_no)}"]
            print("{0}making {1}, please wait till completion".format(counter,Sem_no))
            if ws_sheet.max_row==1 : #adding header row
                ws_sheet.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
            SubCode,Credit,Grade,Sub_Type = data[2:] #initializing variables
            #initializing variables
            subname,ltp,crd = subject_dict[f"{SubCode}"][0],subject_dict[f"{SubCode}"][1],subject_dict[f"{SubCode}"][2]
            #adding data to sheet       
            ws_sheet.append([ws_sheet.max_row,SubCode,subname,ltp,crd,Sub_Type,Grade])
            wb_sheet.save(r'output\\{}.xlsx'.format(Rollno))          
            counter+=1;
    return

def is_sheet_exist(sheet_name,wb_sheet): #function to check the existance of sheet ina workbook
    for sheet in wb_sheet.sheetnames:
        if sheet == sheet_name: #condition for checking the existance
            return 1
    return 0

def get_subject_dictionary(): #fuction to create subject dictionary
    #opening and reading subjects file
    with open("subjects_master.csv", "r") as Subject_data:
        subject_csv_file=csv.reader(Subject_data)
        subject_list=[]
        for data in subject_csv_file:
            subject_list.append(list(data))
        subject_list=subject_list[1:]
        #creating a list of subject data
        #creating a dictionary to store the data
        subject_dict = {}
        length = 0
        #adding data
        for data in subject_list:
            subno = data[0] 
            if subno not in subject_dict: #checking for existant data
                subject_dict[subno] = data[1:]
                length = max(length,len(data[1]))  
    return subject_dict,length

def Spi_calculator(wb_sheet): #function to calculate SPI
    Spi,Credits= [],[]  #initializing lists
    for sheet in wb_sheet.sheetnames[1:]:
        ws_sheet = wb_sheet[sheet]  
        credits = [int(cell.value) for cell in ws_sheet["E"][1:]]
        spi = [grade_to_score[cell.value.strip().strip("*")] for cell in ws_sheet["G"][1:]]
        Credits.append(sum(credits))
        #spi calculation formula
        Spi.append(round(sum([spi[i]*credits[i] for i in range(len(spi))])/sum(credits),2))
    return Spi,Credits

def get_workbook(data): #creating a workbook
    wb_sheet=Workbook()
    ws_sheet=wb_sheet.active
    ws_sheet.title = "Overall"
    wb_sheet.save(f'output\\{data[0]}.xlsx')
    return

def generate_marksheet(): #main function
    Directory =  "output"
    #checking the existence of directory
    if not os.path.exists(Directory):
        os.mkdir(Directory)

    #making a dictionary to store subject data
    subject_dict=[]
    subject_dict,length=get_subject_dictionary()

    #program to make scoresheets
    score_sheet_maker(subject_dict,length)
    
    #opening and reading names-roll file
    with open("names-roll.csv","r") as names_data:
        names_csv = csv.reader(names_data)
        #created a list for names
        names_list=[]
        for data in names_csv:
            names_list.append(list(data))
        names_list=names_list[1:]
        
        #adding data
        for data in names_list:
            Name,Rollno = data[1],data[0] #initializing the variables
            wb_sheet = load_workbook(r'output\\{}.xlsx'.format(Rollno))
            ws_sheet =wb_sheet.active
            #adding data to sheets
            ws_sheet.column_dimensions["A"].width = 30
            ws_sheet.append(["RollNo",Rollno])
            ws_sheet.append(["Name of Student",Name])
            ws_sheet.append(["Branch/Department",Rollno[4:6]])
            #calculating SPI
            Spi,Credits= Spi_calculator(wb_sheet)
            #calculating results
            result_calculate(Spi,ws_sheet,Credits)
            
            wb_sheet.save(r'output\\{}.xlsx'.format(Rollno)) 
    return
generate_marksheet() #calling the function

print('My name is Kanakala Durga Prasanth. My roll no is 1901ee31')