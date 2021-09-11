print('My name is Kanakala Durga Prasanth. My roll no is 1901EE31. PS: please wait 2 mins for compilation of code')

import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
header=['rollno','register_sem','subno','sub_type']
def output_by_subject():
    DIRECTORY = "output_by_subject"
 
    if not os.path.exists(DIRECTORY): #checking the existence of directory
        os.makedirs(DIRECTORY)
    with open('regtable_old.csv','r') as info:
        sinfo=csv.reader(info)
        for data in sinfo:
            del data[4:8]
            del data[2:3]
            if (data[2] =="subno"):continue
            sheetname='{}.xlsx'.format(data[2])
            path='./output_by_subject/'+sheetname
        
            if(os.path.isfile(path)): #adding data to sheet
                 wb=load_workbook(r'output_by_subject\\{}.xlsx'.format(data[2]))
                 xsheet=wb.active
                 xsheet.append(data)
                 wb.save(r'output_by_subject\\{}.xlsx'.format(data[2]))
                 
            else: #creating a new sheet and adding header to sheet 
                    wb=Workbook()
                    xsheet=wb.active
                    xsheet.append(header)
                    xsheet.append(data)
                    wb.save(f'output_by_subject\\{data[2]}.xlsx')
    return
 
def output_individual_roll():
    DIRECTORY = "output_individual_roll"
 
    if not os.path.exists(DIRECTORY): #checking the existence of directory
        os.makedirs(DIRECTORY)
    with open('regtable_old.csv','r') as info:
        sinfo=csv.reader(info)
        for data in sinfo:
            del data[4:8]
            del data[2:3]
            if (data[0] =="rollno"):continue
            sheetname='{}.xlsx'.format(data[0])
            path='./output_individual_roll/'+sheetname
            if(os.path.isfile(path)): #adding data to sheet
                 wb=load_workbook(r'output_individual_roll\\{}.xlsx'.format(data[0]))
                 xsheet=wb.active
                 xsheet.append(data)
                 wb.save(r'output_individual_roll\\{}.xlsx'.format(data[0]))
                 
                    
            else: #creating a new sheet and adding header to sheet 
                    wb=Workbook()
                    xsheet=wb.active
                    xsheet.append(header)
                    xsheet.append(data)
                    wb.save(f'output_individual_roll\\{data[0]}.xlsx')
    return
 
output_by_subject()
output_individual_roll()
