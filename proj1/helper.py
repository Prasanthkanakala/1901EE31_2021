import os
import csv
from numpy import NAN,nan
from PIL.Image import linear_gradient
import pandas as pd
from openpyxl import Workbook
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment,Border,Side, alignment
import mini
path=os.getcwd()
#input=path+'/sample_input/'

#os.chdir(input)

response_file=pd.read_csv("./sample_input/responses.csv")
master_file=pd.read_csv("./sample_input/master_roll.csv")

sample_output_path = "./sample_output/marksheet"
pos=5
neg=1
zero=0
final_score=[]
status=[]

def get_roll_email():
    lst = []
    for index,row in response_file.iterrows():
        lst.append([row["Roll Number"],row["Email address"],row["IITP webmail"]])
    return lst

def generate_marksheet(pos=5,neg=1):
    
    os.makedirs("sample_output",exist_ok = True)
    os.makedirs(sample_output_path,exist_ok = True)
    if(neg <1):
        #neg=-(neg)
        neg=-neg

    #master_file=pd.read_csv("./sample_input/master_roll.csv")
    print(master_file.head())
    master_roll=master_file['roll'].values.tolist()
    master_name=master_file['name'].values.tolist()
    #print(master_roll)


    dict_roll={}
    for i in range(len(master_roll)):
        dict_roll[master_roll[i]]=master_name[i]
        print(dict_roll[master_roll[i]])

    #response_file=pd.read_csv("responses.csv")
    print(response_file.head())
    response_roll=response_file['Roll Number'].values.tolist()
    correct_options=response_file.iloc[0].values.tolist()
    correct_options=correct_options[6:]

    #os.chdir(path)
    #if not os.path.exists('sample_output\marksheet'):
    #    os.mkdir('sample_output\marksheet')

    #status=[]
    #final_score=[]
    no_stu_master=0
    no_of_response=0
    #print(response_roll)
    os.chdir(path)

    alignment_heading=Alignment(horizontal='right',vertical='bottom')
    alignment_content=Alignment(horizontal='left',vertical='bottom')
    alignment_ans=Alignment(horizontal='center',vertical='bottom')
    font_heading=Font(name='Calibri',size=14,bold=False)
    font_content=Font(name='Calibri',size=14,bold=True)
    right_color=Font(color='00FF00',name='Calibri',size=14,bold=False)
    left_color=Font(color='ff0000',name='Calibri',size=14,bold=False)
    give_color=Font(color='0000FF',name='Calibri',size=14,bold=False)
    border_style=Side(border_style='medium',color='000000')

    while(no_stu_master<len(master_roll)):
        wb=Workbook()
        sheet=wb.active
        img=Image('iitp_logo.png')
        alignment_heading=Alignment(horizontal='right',vertical='bottom')
        alignment_content=Alignment(horizontal='left',vertical='bottom')
        alignment_ans=Alignment(horizontal='center',vertical='bottom')
        font_heading=Font(name='Calibri',size=14,bold=False)
        font_content=Font(name='Calibri',size=14,bold=True)
        right_color=Font(color='00FF00',name='Calibri',size=14,bold=False)
        left_color=Font(color='ff0000',name='Calibri',size=14,bold=False)
        give_color=Font(color='0000FF',name='Calibri',size=14,bold=False)
        border_style=Side(border_style='medium',color='000000')
        border=Border(top=border_style,bottom=border_style,left=border_style,right=border_style)
        img.width=610
        img.height=80
        sheet.add_image(img,'A1')
        sheet.column_dimensions['A'].width=17
        sheet.column_dimensions['B'].width=17
        sheet.column_dimensions['C'].width=17
        sheet.column_dimensions['D'].width=17
        sheet.column_dimensions['E'].width=17

        sheet['A6']='Name :'
        sheet['A6'].font=font_heading
        sheet['A6'].alignment=alignment_heading
        
        sheet['D6']='Exam :'
        sheet['D6'].font=font_heading
        sheet['D6'].alignment=alignment_heading
        sheet['E6']='quiz'
        sheet['E6'].font=font_content
        sheet['E6'].alignment=alignment_content
        sheet['A7']='Roll Number :'
        sheet['A7'].font=font_heading
        sheet['A7'].alignment=alignment_heading

        empty=['A9','A10','A11','A12','B9','C9','D9','E9','A15','B15','D15','E15']
        for key in empty:
            sheet[key].font=font_content
            sheet[key].alignment=alignment_ans
            sheet[key].border=border

        lst=['B11','C11','E11','D11','D12','E10']
        for key in lst:
            sheet[key].border=border
        
        mini.alligner(sheet,alignment_ans,font_content,alignment_content,dict_roll,master_roll,no_stu_master,correct_options,pos,neg,zero)

        line=16
        m=0
        while(m<len(correct_options)):
            if(line<=40):
                sheet.cell(row=line,column=2).font=Font(size=12,color='000000FF', name='Calibri')
                sheet.cell(row=line,column=2).alignment=Alignment('center')
                sheet.cell(row=line,column=2).value=correct_options[m]
                sheet.cell(row=line,column=2).border=border

            else:
                sheet.cell(row=line-25,column=5).font=Font(size=12,color='000000FF', name='Calibri')
                sheet.cell(row=line-25,column=5).alignment=Alignment('center')
                sheet.cell(row=line-25,column=5).value=correct_options[m]
                sheet.cell(row=line-25,column=5).border=border

            line+=1
            m+=1

        #set_border(sheet,'A15:B40')
        #set_border(sheet,'D15:E'+'{}'.format(line-25))

        sheet.cell(row=10,column=1).alignment=Alignment('center')
        sheet.cell(row=10,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=10,column=1).value='No.'

        sheet.cell(row=11,column=1).alignment=Alignment('center')
        sheet.cell(row=11,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=11,column=1).value='Marking'

        sheet.cell(row=12,column=1).alignment=Alignment('center')
        sheet.cell(row=12,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=12,column=1).value='Total'

        if(no_of_response<len(response_roll) and master_roll[no_stu_master]==response_roll[no_of_response]):

            stu_options=[]
            stu_options=response_file.iloc[no_of_response].values.tolist()
            stu_options=stu_options[6:]

            k=0
            total=0
            wrong=0
            right=0
            notattempt=0
            while(k<len(correct_options)):
                if(stu_options[k]==correct_options[k]):
                    right+=1
                k+=1

            nonattempt=stu_options.count(nan)
            wrong=len(stu_options)-right-notattempt
            status.append('[{},{},{}]'.format(right,wrong,notattempt))

            total=((right*pos)-(wrong*neg))
            max_marks=len(correct_options)*pos
            final_score.append('{}/{}'.format(total,max_marks))

            #set_border(sheet,'A9:E12')

            sheet.cell(row=10,column=2).font=Font(size=12,color="00008000",name="Calibri")
            sheet.cell(row=10,column=2).alignment=Alignment('center')
            sheet.cell(row=10,column=2).value=str(right)
            sheet.cell(row=10,column=2).border=border

            sheet.cell(row=10,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
            sheet.cell(row=10,column=3).alignment=Alignment('center')
            sheet.cell(row=10,column=3).value=str(wrong)
            sheet.cell(row=10,column=3).border=border

            sheet.cell(row=10,column=4).font=Font(size=12,name="Calibri")
            sheet.cell(row=10,column=4).alignment=Alignment('center')
            sheet.cell(row=10,column=4).value=str(notattempt)
            sheet.cell(row=10,column=4).border=border
            
            sheet.cell(row=12,column=2).font=Font(size=12,color="00008000",name="Calibri")
            sheet.cell(row=12,column=2).alignment=Alignment('center')
            sheet.cell(row=12,column=2).value=str(right*5)
            sheet.cell(row=12,column=2).border=border

            sheet.cell(row=12,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
            sheet.cell(row=12,column=3).alignment=Alignment('center')
            sheet.cell(row=12,column=3).value=str(wrong*(-1))
            sheet.cell(row=12,column=3).border=border

            sheet.cell(row=12,column=5).font=Font(size=12,name="Calibri",color='000000FF')
            sheet.cell(row=12,column=5).alignment=Alignment('center')
            sheet.cell(row=12,column=5).value='{}/{}'.format(total,max_marks)
            sheet.cell(row=12,column=5).border=border


            line=16
            m=0
            mini.maker(correct_options,m,line,stu_options,sheet,border)
            no_of_response+=1
        
        wb.save(f'./sample_output/marksheet/{master_roll[no_stu_master]}.xlsx')
        no_stu_master+=1




def generate_concise_marksheet(mrks,wmrks):
    concise_marksheet = response_file
    crt_opts_list = mini.get_answer()
    if crt_opts_list == []:
        return "Error!!!Answer is not exist in responses"
    last_list,score_af_neg= [],[]
    for index,row in response_file.iterrows():
        curr_opts = [key for key in row][7:]
        result,lst = mini.calculate(mrks,wmrks,crt_opts_list,curr_opts)
        last_list.append(lst)
        score_af_neg.append(str(result)+"/"+str(mrks*len(crt_opts_list)))
    concise_marksheet.insert(loc =6,column ="Score_After_Negative",value =score_af_neg)
    concise_marksheet["Options"] = last_list
    os.makedirs("sample_output",exist_ok = True)
    os.makedirs(sample_output_path,exist_ok = True)
    concise_marksheet.to_csv(sample_output_path+"/concise_marksheet.csv", index=False)
    return  
        
#generate_concise_marksheet()
#generate_marksheet()
#generate_concise_marksheet()
        








    
    





